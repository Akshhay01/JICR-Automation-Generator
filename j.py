import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os

# Constants
DATA_FILE = "DATASET.xlsx"
TEMPLATE_FILE = "TEMPLATE.xlsx"
OUTPUT_FOLDER = "JICR_Reports"
START_ROW = 14
FOOTER_ORIGINAL_START = 54
FOOTER_ORIGINAL_END = 66
MAX_COL = 10

# Ensure files exist
if not os.path.exists(DATA_FILE) or not os.path.exists(TEMPLATE_FILE):
    print("❌ Required files not found.")
    input("Press Enter to exit...")
    exit()

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Load dataset
df = pd.read_excel(DATA_FILE)
df.columns = df.columns.str.strip()

# Identify normalized column names
panchayat_col = [c for c in df.columns if "panchayat" in c.lower()][0]
ward_col = [c for c in df.columns if "ward" in c.lower()][0]

# Ask how many JICR to generate
try:
    total_jicr = int(input("How many JICR do you want to generate? : ").strip())
except ValueError:
    print("❌ Please enter a valid number.")
    input("Press Enter to exit...")
    exit()

for j in range(total_jicr):

    print(f"\n--- JICR {j+1} ---")

    panchayat = input("Enter Panchayat name (case-sensitive): ").strip()
    ward_input = input("Enter comma-separated Ward Nos (e.g., 1,2,3): ").strip()
    wards = [w.strip() for w in ward_input.split(",")]

    # Filter rows
    filtered_df = df[(df[panchayat_col] == panchayat) & (df[ward_col].astype(str).isin(wards))]
    final_rows = pd.DataFrame()

    for ward in wards:
        ward_rows = filtered_df[filtered_df[ward_col].astype(str) == ward].head(10)
        final_rows = pd.concat([final_rows, ward_rows])

    if final_rows.empty:
        print("❌ No matching data found for this Panchayat.")
        continue

    # Load template fresh each time
    wb = load_workbook(TEMPLATE_FILE)
    template_wb = load_workbook(TEMPLATE_FILE)

    # ---------------- FORM 10 ----------------
    ws_form = wb["FORM 10"]
    ws_template_form = template_wb["FORM 10"]

    for row in ws_form.iter_rows(min_row=START_ROW, max_row=200, max_col=MAX_COL):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    ws_form.delete_rows(FOOTER_ORIGINAL_START, FOOTER_ORIGINAL_END - FOOTER_ORIGINAL_START + 1)

    for idx, (_, row) in enumerate(final_rows.iterrows()):
        r = START_ROW + idx
        ws_form.cell(row=r, column=1, value=idx + 1)
        ws_form.cell(row=r, column=2, value=row["District"])
        ws_form.cell(row=r, column=3, value=row["Block"])
        ws_form.cell(row=r, column=4, value=row["Panchayat"])
        ws_form.cell(row=r, column=5, value=row["Ward No"])
        ws_form.cell(row=r, column=6, value=row["Pole Id"])
        ws_form.cell(row=r, column=7, value=row["HOUSE HOLDER NAME"])

        for col in range(1, MAX_COL + 1):
            src_cell = ws_template_form.cell(row=14, column=col)
            tgt_cell = ws_form.cell(row=r, column=col)
            if src_cell.has_style:
                tgt_cell._style = src_cell._style

    data_end_row = START_ROW + len(final_rows)
    footer_target_start = data_end_row + 1
    ws_form.insert_rows(footer_target_start, FOOTER_ORIGINAL_END - FOOTER_ORIGINAL_START + 1)

    for i in range(FOOTER_ORIGINAL_END - FOOTER_ORIGINAL_START + 1):
        src_row = FOOTER_ORIGINAL_START + i
        tgt_row = footer_target_start + i
        for col in range(1, MAX_COL + 1):
            src = ws_template_form.cell(row=src_row, column=col)
            tgt = ws_form.cell(row=tgt_row, column=col)
            tgt.value = src.value
            if src.has_style:
                tgt._style = src._style

    # ---------------- LUMINARY ----------------
    lum_ws = wb["LUMINARY"]
    lum_template = template_wb["LUMINARY"]
    lum_ws.delete_rows(3, lum_ws.max_row - 2)

    for idx, (_, row) in enumerate(final_rows.iterrows()):
        r = 3 + idx
        for col in range(1, 14):
            tgt = lum_ws.cell(row=r, column=col)
            src = lum_template.cell(row=3, column=col)

            if col == 1:
                tgt.value = idx + 1
            elif col == 2:
                tgt.value = row["Pole Id"]
            elif col == 3:
                tgt.value = row["Luminary Serial No"]
            elif col == 11:
                tgt.value = row["Ward No"]
            else:
                tgt.value = src.value

            if src.has_style:
                tgt._style = src._style

    # ---------------- BATTERY ----------------
    bat_ws = wb["BATTERY"]
    bat_template = template_wb["BATTERY"]
    bat_ws.delete_rows(3, bat_ws.max_row - 2)

    for idx, (_, row) in enumerate(final_rows.iterrows()):
        r = 3 + idx
        for col in range(1, 9):
            tgt = bat_ws.cell(row=r, column=col)
            src = bat_template.cell(row=3, column=col)

            if col == 1:
                tgt.value = idx + 1
            elif col == 2:
                tgt.value = row["Pole Id"]
            elif col == 4:
                tgt.value = row["Battery Serial No"]
            elif col == 8:
                tgt.value = row["Ward No"]
            else:
                tgt.value = src.value

            if src.has_style:
                tgt._style = src._style

    # ---------------- SOLAR ----------------
    solar_ws = wb["SOLAR"]
    solar_template = template_wb["SOLAR"]
    solar_ws.delete_rows(3, solar_ws.max_row - 2)

    for idx, (_, row) in enumerate(final_rows.iterrows()):
        r = 3 + idx
        for col in range(1, 12):
            tgt = solar_ws.cell(row=r, column=col)
            src = solar_template.cell(row=3, column=col)

            if col == 1:
                tgt.value = idx + 1
            elif col == 2:
                tgt.value = row["Pole Id"]
            elif col == 4:
                tgt.value = row["Solar Panel Serial No"]
            elif col == 11:
                tgt.value = row["Ward No"]
            else:
                tgt.value = src.value

            if src.has_style:
                tgt._style = src._style

    # ---------------- DETAILS ----------------
    details_ws = wb["DETAILS"]
    details_template = template_wb["DETAILS"]
    details_ws.delete_rows(3, details_ws.max_row - 2)

    for idx, (_, row) in enumerate(final_rows.iterrows()):
        r = 3 + idx
        values = [
            idx + 1,
            row["District"],
            row["Block"],
            row["Panchayat"],
            row["Ward No"],
            row["Pole Id"],
            row["LAT."],
            row["LON."],
            row["HOUSE HOLDER NAME"],
            ""
        ]

        for col in range(1, 11):
            tgt = details_ws.cell(row=r, column=col)
            src = details_template.cell(row=3, column=col)
            tgt.value = values[col - 1]
            if src.has_style:
                tgt._style = src._style
    

    footer_start = 3 + len(final_rows)
    details_ws.insert_rows(footer_start, 7)
    for i in range(7):
        src_row = 43 + i
        tgt_row = footer_start + i
        for col in range(1, 11):
            src = details_template.cell(row=src_row, column=col)
            tgt = details_ws.cell(row=tgt_row, column=col)
            tgt.value = src.value
            if src.has_style:
               tgt._style = src._style


    # Save workbook
    output_name = f"JICR_{panchayat}_Wards_{'-'.join(wards)}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)
    wb.save(output_path)

    print(f"✅ JICR generated successfully for {panchayat}")

print("\n🎉 All JICR files generated successfully.")
input("Press Enter to exit...")
